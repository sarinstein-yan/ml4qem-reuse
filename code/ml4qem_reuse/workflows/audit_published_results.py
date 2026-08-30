#!/usr/bin/env python3
"""Recompute published aggregates from the distributed result tables."""

from __future__ import annotations

import argparse
import json
import math
import pickle
import re
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook

from ml4qem_reuse.workflows._paths import workspace_root


PROJECT = workspace_root()
PAPER_RESULTS = (
    PROJECT / "upstream/snapshots/qiskit-community-ml-qem-9776e1b/docs/paper_figures"
)
OFFICIAL_SOURCE = (
    PROJECT
    / "refs/Liao et al. - 2024 - Machine learning for practical quantum error mitigation - Source data.xlsx"
)
REPOSITORY_SOURCE = PAPER_RESULTS / "ML-QEM Source data.xlsx"


def _load_pickle(name: str) -> Any:
    with (PAPER_RESULTS / name).open("rb") as stream:
        return pickle.load(stream)


def _extract_series_value(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    matches = re.findall(r"[-+]?(?:\d+\.\d+|\d+)(?:[eE][-+]?\d+)?", str(value))
    if len(matches) < 2:
        raise ValueError(f"cannot parse pandas-series cell: {value!r}")
    return float(matches[-1])


def _difference_record(source: float, recomputed: float) -> Dict[str, float]:
    return {
        "source": float(source),
        "recomputed": float(recomputed),
        "signed_difference": float(recomputed - source),
        "absolute_difference": float(abs(recomputed - source)),
    }


def _figure2(workbook: Any) -> Dict[str, Any]:
    source = {row[0]: float(row[1]) for row in workbook["Figure 2"].iter_rows(values_only=True)}
    frame = _load_pickle("random_circuits.pk")["df"]
    mapping = {
        "Unmitigated": "L2_noisy",
        "ZNE": "L2_zne",
        "OLS": "L2_ols_full",
        "RF": "L2_rfr_list",
        "MLP": "L2_mlp",
        "GNN": "L2_gnn",
    }
    comparisons = {
        method: _difference_record(source[method], frame[column].mean())
        for method, column in mapping.items()
    }
    return {
        "caption_n": 2000,
        "distributed_result_rows": int(len(frame)),
        "rows_by_depth": {str(key): int(value) for key, value in frame.groupby("step").size().items()},
        "missing_relative_to_caption": 2000 - int(len(frame)),
        "comparisons": comparisons,
        "max_absolute_difference": max(value["absolute_difference"] for value in comparisons.values()),
    }


def _figure3(workbook: Any) -> Dict[str, Any]:
    sheet = workbook["Figure 3"]
    result_files = (
        "no_readout_over_depths.pk",
        "incoherent_over_depths.pk",
        "coherent_over_depths.pk",
    )
    columns = (1, 3, 5)
    blocks = {
        "L2_noisy": range(3, 33),
        "L2_zne": range(35, 65),
        "L2_rfr_list": range(67, 97),
    }
    cells = []
    for column, filename in zip(columns, result_files):
        frame = _load_pickle(filename)["df_mean"]
        for metric, rows in blocks.items():
            for depth, row in enumerate(rows):
                source = _extract_series_value(sheet.cell(row, column).value)
                recomputed = float(frame.loc[depth, metric])
                cells.append(
                    {
                        "result_file": filename,
                        "metric": metric,
                        "depth": depth,
                        **_difference_record(source, recomputed),
                    }
                )
    return {
        "cells_compared": len(cells),
        "source_precision_decimal_places": 6,
        "max_absolute_difference": max(cell["absolute_difference"] for cell in cells),
        "cells_with_difference_above_5e-7": sum(
            cell["absolute_difference"] > 5e-7 for cell in cells
        ),
        "comparisons": cells,
    }


def _figure4(workbook: Any) -> Dict[str, Any]:
    sheet = workbook["Figure 4"]
    frame = _load_pickle("hardware_over_depth.pk")["df_mean"]
    blocks = {
        "L2_noisy": range(2, 12),
        "L2_rfr_list": range(14, 24),
        "L2_zne": range(26, 36),
    }
    cells = []
    for metric, rows in blocks.items():
        for depth, row in enumerate(rows):
            source = _extract_series_value(sheet.cell(row, 1).value)
            recomputed = float(frame.loc[depth, metric])
            cells.append({"metric": metric, "depth": depth, **_difference_record(source, recomputed)})
    return {
        "selected_result_file": "hardware_over_depth.pk",
        "ambiguous_alternative_result_file": "hardware_over_depths.pk",
        "cells_compared": len(cells),
        "source_precision_decimal_places": 6,
        "max_absolute_difference": max(cell["absolute_difference"] for cell in cells),
        "cells_with_difference_above_5e-7": sum(
            cell["absolute_difference"] > 5e-7 for cell in cells
        ),
        "comparisons": cells,
    }


def _cell_profile(workbook: Any) -> Dict[str, Any]:
    result = {}
    for sheet in workbook.worksheets:
        nonempty = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        numeric = [
            value
            for value in nonempty
            if isinstance(value, (int, float))
            and not isinstance(value, bool)
            and math.isfinite(float(value))
        ]
        numeric_like_strings = []
        for value in nonempty:
            if not isinstance(value, str):
                continue
            cleaned = value.strip().strip("[],")
            try:
                float(cleaned)
                numeric_like_strings.append(value)
            except ValueError:
                pass
        result[sheet.title] = {
            "nonempty_cells": len(nonempty),
            "native_numeric_cells": len(numeric),
            "numeric_like_string_cells": len(numeric_like_strings),
        }
    return result


def _workbook_differences(first: Any, second: Any) -> list[Dict[str, Any]]:
    differences = []
    for name in first.sheetnames:
        left = first[name]
        right = second[name]
        for row in range(1, max(left.max_row, right.max_row) + 1):
            for column in range(1, max(left.max_column, right.max_column) + 1):
                first_value = left.cell(row, column).value
                second_value = right.cell(row, column).value
                if first_value != second_value:
                    differences.append(
                        {
                            "sheet": name,
                            "row": row,
                            "column": column,
                            "official": first_value,
                            "repository": second_value,
                        }
                    )
    return differences


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    official = load_workbook(OFFICIAL_SOURCE, data_only=True)
    repository = load_workbook(REPOSITORY_SOURCE, data_only=True)
    workbook_differences = _workbook_differences(official, repository)
    result = {
        "schema_version": 1,
        "artifact_commit": "9776e1b",
        "official_source_workbook": str(OFFICIAL_SOURCE.relative_to(PROJECT)),
        "repository_source_workbook": str(REPOSITORY_SOURCE.relative_to(PROJECT)),
        "figure2": _figure2(official),
        "figure3": _figure3(official),
        "figure4": _figure4(official),
        "official_source_cell_profile": _cell_profile(official),
        "official_repository_workbook_difference_count": len(workbook_differences),
        "official_repository_workbook_differences": workbook_differences,
        "hardware_jobs_submitted": 0,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w") as stream:
        json.dump(result, stream, indent=2, sort_keys=True)
        stream.write("\n")
    summary = {
        "figure2": {key: value for key, value in result["figure2"].items() if key != "comparisons"},
        "figure3": {key: value for key, value in result["figure3"].items() if key != "comparisons"},
        "figure4": {key: value for key, value in result["figure4"].items() if key != "comparisons"},
        "workbook_difference_count": len(workbook_differences),
    }
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
